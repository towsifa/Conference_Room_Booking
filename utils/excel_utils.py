import logging

import openpyxl


def write_test_data(file_path, sheet_name, row_number, column_number, actual_cost):
    """
    Writes the actual cost into the Excel file at the specified row.

    :param column_number:
    :param file_path: Path to the Excel file.
    :param sheet_name: Name of the sheet to write to.
    :param row_number: The row number to update (starting from 2 for the first data row).
    :param actual_cost: The actual cost to write.
    """
    try:
        # Load the workbook and select the sheet
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        # Write the actual cost
        sheet.cell(row=row_number, column=column_number).value = actual_cost

        # Save the workbook
        workbook.save(file_path)
        print(f"Successfully wrote actual cost to row {row_number}.")

    except Exception as e:
        print(f"Error writing to Excel: {e}")


def read_test_data(file_path, sheet_name):
    """
    Reads test data from an Excel file and ensures correct formatting of date and time.
    :param file_path: Path to the Excel file.
    :param sheet_name: Name of the sheet containing the data.
    :return: List of tuples with test data.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        data = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Extract components
            room_type = row[0]

            start_month = str(row[1]).zfill(2)  # Ensure month is zero-padded
            start_day = str(row[2]).zfill(2)  # Ensure day is zero-padded
            start_year = str(row[3])  # Year as string
            start_hour = str(row[4]).zfill(2)  # Ensure hour is zero-padded
            start_minute = str(row[5]).zfill(2)  # Ensure minute is zero-padded
            start_time = row[6]  # AM/PM value (optional)

            # End date-time components
            end_month = str(row[7]).zfill(2)
            end_day = str(row[8]).zfill(2)
            end_year = str(row[9])
            end_hour = str(row[10]).zfill(2)
            end_minute = str(row[11]).zfill(2)
            end_time = row[12]

            expected_cost = row[14]

            # Append formatted data
            data.append((room_type, start_month, start_day, start_year, start_hour, start_minute, start_time, end_month,
                         end_day, end_year, end_hour, end_minute, end_time, expected_cost))
        return data

    except Exception as e:
        logging.error(f"Error reading test data: {e}")
        return []